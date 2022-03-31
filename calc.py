import openpyxl
import os


def commandlist():
    path = "data/commands.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    message = "Lista moich komend:\n"
    for row in sheet_obj.iter_rows():
        c1 = row[0].value
        c2 = row[1].value
        c3 = row[2].value
        message += (c1 + " - " + c2 + ", argumenty: " + c3 + "\n")
    wb_obj.save(path)
    return message


def new_user_points(p):
    path = "data/points.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for row in sheet_obj.rows:
        for cell in row:
            if cell.value == p:
                message = "Już zliczam Twoje punkty!"
                return message
    row_count = int(sheet_obj.max_row) + 1
    cell1 = sheet_obj.cell(row=row_count, column=1)
    cell1.value = p
    cell2 = sheet_obj.cell(row=row_count, column=3)
    cell2.value = 0
    message = "Od teraz punkty użytkownika " + p + "' są już zliczane!"
    wb_obj.save(path)
    return message


def adding_points(p):
    path = "data/points.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for row in sheet_obj.rows:
        for cell in row:
            if cell.value == p:
                cell1 = sheet_obj.cell(row=cell.row, column=3)
                c1 = cell1.value
                print(c1)
                c1 += 1
                cell1.value = c1
                print(cell1.value)
                print(type(c1))
            break
    wb_obj.save(path)


def get_points(p):
    path = "data/points.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for row in sheet_obj.rows:
        for cell in row:
            print(cell.value)
            if p == cell.value:
                cell1 = sheet_obj.cell(row=cell.row, column=3)
                c1 = cell1.value
                message = "Amount of points for " + p + " is: " + str(c1)
                wb_obj.save(path)
                return message

    message = "Ten użytkownik nie ma jeszcze zliczanych punktów!"
    wb_obj.save(path)
    return message


def track_goals(dire):
    path = "data/"+dire+"/goals.xlsx"
    check = os.path.exists(path)
    if check:
        message = "Już zapisuję Twoje cele!"
        return message
    wb = openpyxl.Workbook()
    wb.save(path)
    message = "Od teraz zapisuję Twoje cele!"
    return message


def add_goal(dire, goal):
    path = "data/"+dire+"/goals.xlsx"
    check = os.path.exists(path)
    if not check:
        message = "Nie zapisuję jeszcze Twoich celów!"
        return message
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    row_count = int(sheet_obj.max_row)+1
    cell1 = sheet_obj.cell(row=row_count, column=1)
    cell1.value = goal
    cell2 = sheet_obj.cell(row=row_count, column=2)
    cell2.value = "N"
    message = "Dodałem nowy cel!"
    wb_obj.save(path)
    return message


def list_goals(dire):
    path = "data/" + dire + "/goals.xlsx"
    check = os.path.exists(path)
    if not check:
        message = "Nie zapisuję jeszcze Twoich celów!"
        return message
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    message = "Cele:\n"
    for index, row in enumerate(sheet_obj.iter_rows()):
        c1 = row[0].value
        c2 = row[1].value
        if c1 is None:
            message += ("\n")
            continue
        if c2 == "N":
            message += (str(index)+". "+c1 + "   Ukończone: Nie"+"\n")
        else:
            message += (str(index)+". "+c1 + "   Ukończone: Tak"+"\n")
    wb_obj.save(path)
    if message == "Cele:\n":
        message = "Nie masz jeszcze żadnych celów - dodaj je!"
    return message


def finish_goal(dire, number):
    if True:
        try:
            number = int(number)
        except ValueError:
            message = "Nie podano numeru celu tylko napis!"
            code = 1
            return [message, code]
    print(type(number))
    path = "data/" + dire + "/goals.xlsx"
    check = os.path.exists(path)
    if not check:
        message = "Nie zapisuję jeszcze Twoich celów!"
        code = 1
        return [message, code]
    if number == 0:
        message = "Podaj właściwy numer celu który ukończono!"
        code = 1
        return [message, code]
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for index, row in enumerate(sheet_obj.iter_rows()):
        c1 = row[1].value
        print(c1)
        if c1 == "N" and index == number:
            message = "Gratulacje! Dostajesz punkt za ukończenie celu!"
            code = 0
            row[1].value = "T"
            print(row[1].value)
            wb_obj.save(path)
            return [message, code]
        else:
            code = 1
            message = "Nie mam celu o takim numerze lub został on już ukończony!"

    return [message, code]