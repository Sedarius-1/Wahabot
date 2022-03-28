import random
import discord
import openpyxl
import os

client = discord.Client()


def commandlist():
    path = "commands.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    message = "Lista moich komend:\n"
    for row in sheet_obj.iter_rows():
        c1 = row[0].value
        c2 = row[1].value
        c3 = row[2].value
        message += (c1 + " - " + c2 + ", argumenty: " + c3 + "\n")
    #print(message)
    wb_obj.save(path)
    return message

def make_user_database(p):
    path = (str(p)[:5])
    print(path)
    isExist = os.path.exists("data/"+path)
    print(isExist)
    if isExist:
        message = "Masz już swoją bazę, nie mogę utworzyć drugiej!"
    if not isExist:
        os.makedirs("data/"+path)
        message = "Utworzyłem dla Ciebie własną bazę - możesz tworzyć już swoje listy!"
    return message


def check_user_database(p):
    path = (str(p)[:5])
    isExist = os.path.exists("data/" + path)
    if not isExist:
        message = "Nie widzę Twojej bazy - przed działaniem musisz ją utworzyć komendą .makebase"
        return message
    if isExist:
        return path


def checklist(dire, p):
    print(type(p))
    if p == None:
        p = 'main'
    print("data/"+dire+'/'+p+'.txt')
    isExist = os.path.exists("data/"+dire+'/'+p+'.txt')
    if not isExist:
        message = "Nie ma takiej listy - musisz ją utworzyć komendą .makelist"
        code = 1
    if isExist:
        message = "Widzę taką listę - zabieram się do działania!"
        code = 0
    return [message, code]


def idea_checklist(path):
    p1='data/'+path+'/main.txt'
    print(p1)
    isExist1 = os.path.exists('data/'+path+'/main.txt')
    if not isExist1:
        message = "Nie ma pliku nazwanego main - nie mogę losować! :("
        code = 1
        return [message, code]
    for filename in os.listdir("data/"+path+"/"):
        if filename != 'main.txt':
            message = "Pliki na miejscu - zabieram się do losowania! :)"
            code = 0
            return [message, code]
    message = "Nie ma pliku nienazwanego main - nie mogę losować! :("
    code = 1
    return [message, code]


def add_list(dire, p):
    a=checklist(dire,p)[1]
    if a == 0:
        message = "Taka lista już istnieje - nadpisz ją używając odpowiedniej komendy!"
        return message
    else:
        with open("data/"+dire+"/"+p+'.txt', 'a') as f1:
            f1.write('')
        f1.close()
        message = "Nie znalazłem takiej listy więc już ją tworzę!"
        with open("data/"+dire+"/main.txt", 'a') as f2:
            f2.write(p+'\n')
        f2.close()
        return message


def activchoice(p, dire):
    with open("data/"+dire+'/'+p + '.txt') as f1:
        lines1 = f1.readlines()
    f1.close()
    z=len(lines1) - 1
    if(z)<0:
        z=0
    n: int = random.randint(0, z)
    print("index:"+str(n))
    print("wylosowane:"+str(lines1[n]))
    nazwa = lines1[n].strip()
    linie = n
    #print(nazwa, '\n', linie)
    return [nazwa, linie]
    # return lines1[n].strip(), n


def emptycheck(dire):
    l1 = activchoice("main", dire)
    l2 = activchoice(l1[0], dire)
    print("l2[0]:"+l2[0])
    if l2[0] == 'p':
        code = 1
    else:
        code = 0
    print(code)
    return [l1, l2, code]


def construct_message(name, dire):

    lista = emptycheck(dire)
    while lista[2] != 0:
        lista = emptycheck(dire)
    p = "Plan na dziś dla " + name + " - "
    constructed_message = p + lista[0][0] + ": " + lista[1][0]
    return constructed_message


def get_list(name):
    with open(name + '.txt') as f1:
        lines1 = f1.readlines()
    f1.close()
    return lines1


def parse_multiple_into_one(amount, li):
    onemes = "List:\n"
    for i in range(amount):
        onemes += (str(i + 1) + ". " + li[i])
    return onemes


def add_to_list(name, addon):
    with open(name + '.txt') as f1:
        lines1 = f1.readlines()
    f1.close()
    if lines1[0] == "p":
        with open(name + '.txt', 'w') as f1:
            f1.write(addon+'\n')
        f1.close()
    else:
        with open(name + '.txt', 'a') as f1:
            f1.write(addon+'\n')
        f1.close()


def new_user_points(p):
    path = "points.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for row in sheet_obj.rows:
        for cell in row:
            if cell.value == p:
                message = "Już zliczam Twoje punkty!"
                return message
    row_count = int(sheet_obj.max_row) + 1
    cell1 = sheet_obj.cell(row=row_count, column=1)
    cell1.value = p
    cell2 = sheet_obj.cell(row=row_count, column=3)
    cell2.value = 0
    message = "Od teraz punkty użytkownika " + p + "' są już zliczane!"
    wb_obj.save(path)
    return message


def adding_points(p):
    path = "points.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for row in sheet_obj.rows:
        for cell in row:
            if cell.value == p:
                cell1 = sheet_obj.cell(row=cell.row, column=3)
                c1 = cell1.value
                print(c1)
                c1 += 1
                cell1.value = c1
                print(cell1.value)
                print(type(c1))
            break
    wb_obj.save(path)


def get_points(p):
    path = "points.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    for row in sheet_obj.rows:
        for cell in row:
            print(cell.value)
            if p == cell.value:
                cell1 = sheet_obj.cell(row=cell.row, column=3)
                c1 = cell1.value
                #print(c1)
                message = "Amount of points for " + p + " is: " + str(c1)
                wb_obj.save(path)
                return message

    message = "Ten użytkownik nie ma jeszcze zliczanych punktów!"
    wb_obj.save(path)
    return message





